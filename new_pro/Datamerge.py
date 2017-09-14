import os
import re
import random
import time,datetime
import numpy as np
from openpyxl import load_workbook,Workbook
import profilo_manage.new_pro.constant_def as NC
nm_stocks = NC.NUM_STOCKS
file_fromat = NC.FILE_FORMAT
start_row = NC.TITLE_START_ROW
feature_list = NC.FEATURES

class Data(object):
    #
    #
    #处理txt文件

    def __init__(self,read_file_path,write_file_path,NUM_STOCKS=nm_stocks):
        '''
        :param read_file_path:  读取文件的地址，到文件夹为止，eg：r'C:\\Users\\wade\\Desktop\\txtfile'
        :param write_file_path: 写入文件的地址，到文件夹为止，eg:r'C:\\Users\\wade\\Desktop\\txtfile'
        :param NUM_STOCKS:  需要生成excel文件的股票数量，每个文件夹中可以放置大于此数量的txt文件
        :param sheet_name:  excel表单名字，生成文件的名字为 r'sheet_mane.xlsx'
        '''
        self.read_file_path = read_file_path
        self.write_file_path = write_file_path
        self.nm_stocks = NUM_STOCKS
        self.sheet_name = 'result_'+ time.strftime("%Y_%m_%d_%H_%M_%S")
        self.day_nums = 0  #excel表中有效的天数
        self.file_iters = self.get_file_iters()
        self.full_lines = self.get_full_lines()
        self.csv_file = self.write_to_excel()

    #
    #
    #获取文件迭代器
    #1.先判断文件夹中文件是否有不小于需要的最小文件数
    #2.去掉文件中不满足要求格式的文件， 此程序目前只能读取txt文件
    #3.判断剩下的txt文本数是否满足需要的文件数
    def get_file_iters(self):    #read txt file
        file_lists = os.listdir(self.read_file_path)

        if len(file_lists) < self.nm_stocks: #判断文件夹文件数量
            print('file is not enough')
            exit(-1)
        for file_name in file_lists:
            if file_name[-len(file_fromat):] != file_fromat:
                file_lists.remove(file_name)
        file_lists = random.sample(file_lists,self.nm_stocks)

        if len(file_lists)!=self.nm_stocks: #判断文件夹文件数量
            print('wrong file numbers')
            exit(-1)

        self.files = file_lists

        file_iters = []
        for file_name in file_lists:
            file = self.read_file_path + '\\' + file_name
            f = open(file)
            file_iters.append(f)
        if len(file_iters)==0:
            print('no file')
            exit(-1)
        else:
            return file_iters

    def is_datetime(self,str):
        #print('str',str)
        pattern = re.compile(r'(.*)(\d{4}/\d{1,2}/\d{1,2})')
        match = pattern.match(str)
        if match:
            return True
        else:
            return False

    def get_line_start_with_datetime(self,iter): #得到一个文件的下一个第一个元素是日期的行
        try:
            line = iter.__next__()
            #print(line)
            line = line.split()
            # line = iter.__next__().split()    #返回的是一个字符串，用tab分隔开的，需要split
            # print(line)
            # print(self.is_datetime(line[0]))
            while(not self.is_datetime(line[0])):
                line = iter.__next__().split()
            return line
        except StopIteration:
            print('stop')

            raise

    def change_to_datetime(self,str):  #将字符串转换成datetime，方便比较大小
        year_month_day = str.split(r'/')
        if len(year_month_day[1]) == 1:
            year_month_day[1] = '0' + year_month_day[1]

        if len(year_month_day[2]) == 1:
            year_month_day[2] = '0' + year_month_day[2]

        year_month_day = year_month_day[0] + year_month_day[1] + year_month_day[2]
        date = time.strptime(year_month_day, '%Y%m%d')
        date_time = datetime.datetime(date[0], date[1], date[2])
        date_time = date_time.strftime('%Y/%m/%d')

        return date_time

    def get_a_right_line(self):
        try:
            is_right = False #判断10行数据是否日期相同
            index_not_max_datetime = [i for i in range(self.nm_stocks)] #不是最大日期的文件索引，初始时，是所有文件都不是
            nm_stocks_lines = [0 for i in range(self.nm_stocks)]
            while(not is_right):

                for index in index_not_max_datetime:     #添加不是maxdate的数据，但不一定日期一样
                    line = self.get_line_start_with_datetime(self.file_iters[index])
                    nm_stocks_lines[index] = line

                datetimes = []
                for stock_index in range(self.nm_stocks):
                    datetimes.append(nm_stocks_lines[stock_index][0])
                max_datetime = max(datetimes)
                min_datetime = min(datetimes)
                if min_datetime==max_datetime:
                    is_right = True
                    break
                index_not_max_datetime = []
                for index in range(len(datetimes)):
                    if not datetimes[index] == max_datetime:
                        index_not_max_datetime.append(index)



            return nm_stocks_lines
        except StopIteration:
            raise

    # 10个文件，每一行是一天，存放在一个list里，10列,
    # full_lines是十个文件里的所有数据
    def get_full_lines(self):
        full_lines = []
        while(True):
            try:
                line = self.get_a_right_line()
                full_lines.append(line)

            except StopIteration:
                print('all done')
                break
        # print(full_lines)
        return full_lines


    #
    #将一天的数据写进excel表格中，数据形式如下
    # [['2017/09/08' 'open' '45.65' '211.19']
    #  ['2017/09/08' 'high' '45.90' '214.69']
    #  ['2017/09/08' 'low' '45.47' '210.09']
    #  ['2017/09/08' 'close' '45.77' '212.94']
    #  ['2017/09/08' 'volumn' '40244300' '26046900']
    #  ['2017/09/08' 'amount' '418564096.00' '237605968.00']]
    def write_a_day_data(self,ws,datas,rows):
        '''
        :param ws: excel work sheet
        :param datas: one day datas of all stocks
        :param rows: which row to write
        :return: the first row of the next day datas
        '''
        for line in datas:
            ws.cell(row=rows,column=NC.START_COLUMN_DATE,value=line[0])
            ws.cell(row=rows,column=NC.START_COLUMN_DATE+1,value=line[1])
            for i in range(2,2+self.nm_stocks):
                # print('i',i)
                # print('column',NC.START_COLUMN_DATE+i)
                # print('value',float(line[i]))
                ws.cell(row=rows,column=NC.START_COLUMN_DATE+i,value=float(line[i]))
            rows += 1
        return rows


    def write_to_excel(self):
        ###########
        #write title
        wb = Workbook()
        ws = wb.create_sheet(title=self.sheet_name,index=0)
        save_path = self.write_file_path +r'\\'+self.sheet_name+ r'.xlsx'
        ##############
        #write titles
        ws.cell(row=start_row,column=NC.START_COLUMN_DATE,value='date')
        ws.cell(row=start_row,column=NC.START_COLUMN_DATE+1,value='features')
        for i in range(self.nm_stocks):
            ws.cell(row=start_row,column=NC.START_COLUMN_DATE+2+i,value=self.files[i][:-len(file_fromat)-1])
        start_row_writing = start_row + 1 #开始写每行具体数据的其实行号
        for line in self.full_lines:
            self.day_nums += 1
            date = line[0][0] #当天的日期
            dates = [] #列表，保存了features数量的日期
            for i in range(len(feature_list)):
                dates.append(date)
            line = [row[1:] for row in line]
            line.insert(0,feature_list) #插入特征字符串，'open','close',....
            line.insert(0,dates)    #插入日期
            #
            line = np.array(line).T #将line的形式转换成(num_features,num_stocks+2), 2 is datetime and feature`s name
            # print(line)
            start_row_writing = self.write_a_day_data(ws,line,start_row_writing)

        ws.cell(row=1,column=1,value=self.day_nums)
        wb.save(save_path)
        print(save_path)


if __name__ == '__main__':
    read_path = r'C:\Users\wade\Desktop\txtfile'
    data = Data(read_path,read_path)

    # file_iter = data.get_file_iters()
    # data.get_full_lines()
    # print(data.change_to_datetime('2014/5/17'))

    #print(data.full_lines)

    #
    #测试日期是否都是相同的
    # for line in data.full_lines:
    #     date = line[0][0]
    #     for i in range(len(line)):
    #         if line[i][0] != date:
    #             print('fail')



